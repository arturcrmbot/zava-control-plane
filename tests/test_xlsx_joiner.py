from pathlib import Path
import openpyxl
import pytest
from helpers.xlsx_build.joiner import build_questionnaire_xlsx


def test_build_joins_template_and_answers(tmp_path, fixtures_dir):
    out = tmp_path / "out.xlsx"
    report = build_questionnaire_xlsx(
        template_xlsx=fixtures_dir / "mini-template.xlsx",
        answers_dir=fixtures_dir / "mini-answers",
        output_xlsx=out,
    )
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Instructions", "Questionnaire"]
    ws = wb["Questionnaire"]
    # header + 4 rows (the mini fixtures now include 10.1)
    assert ws.max_row == 5
    # 10 columns
    assert ws.max_column == 10
    # header row
    header = [cell.value for cell in ws[1]]
    assert header == ["Ref", "Section", "Subsection", "Question", "MoSCoW",
                      "Status", "Response", "Key Technologies", "POC Demo", "Reference"]
    # first data row is Ref 1.1 with our columns appended
    row = [cell.value for cell in ws[2]]
    assert row[0] == "1.1"
    assert row[5] == "Can do today"
    assert row[6] == "Test response one"


def test_build_reports_no_mismatches_for_clean_input(tmp_path, fixtures_dir):
    out = tmp_path / "out.xlsx"
    report = build_questionnaire_xlsx(
        template_xlsx=fixtures_dir / "mini-template.xlsx",
        answers_dir=fixtures_dir / "mini-answers",
        output_xlsx=out,
    )
    assert report.missing_refs == []
    assert report.extra_refs == []
    assert report.question_text_mismatches == []
    assert report.row_count == 4


def test_build_flags_missing_ref(tmp_path, fixtures_dir):
    # Create a broken answers dir lacking Ref 2.1 and 10.1 (only 01-test.csv present)
    broken = tmp_path / "broken_answers"
    broken.mkdir()
    src = (fixtures_dir / "mini-answers" / "01-test.csv").read_text(encoding="utf-8")
    (broken / "01-test.csv").write_text(src, encoding="utf-8")

    out = tmp_path / "out.xlsx"
    report = build_questionnaire_xlsx(
        template_xlsx=fixtures_dir / "mini-template.xlsx",
        answers_dir=broken,
        output_xlsx=out,
    )
    assert "2.1" in report.missing_refs
    assert "10.1" in report.missing_refs


def test_duplicate_template_ref_matched_by_question_text(tmp_path, fixtures_dir):
    """WPP's real template has a typo: two rows both labelled "8.1"; the second
    should be "8.10". Our CSVs use "8.10" correctly. The builder must match
    the CSV's "8.10" to the template's second "8.1" row via question text.

    Build a mini template with two "1.1" rows (different question text) and two
    matching CSV answers (one with ref "1.1" matching first question, one with
    ref "5.99" matching second question). The builder should join them by
    question text when template Ref is ambiguous and preserve the template's
    Ref column values in the output.
    """
    # Build a local template with duplicate "1.1" entries
    template = tmp_path / "dup_template.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    instr = wb.create_sheet("Instructions")
    instr.append(["#", "Instruction"])
    q = wb.create_sheet("Questionnaire")
    q.append(["Ref", "Section", "Subsection", "Question", "MoSCoW"])
    q.append(["1.1", "S", "Sub", "First question labelled 1.1", "Must"])
    q.append(["1.1", "S", "Sub", "Second question mislabelled (should be higher ref)", "Must"])
    wb.save(template)

    answers = tmp_path / "dup_answers"
    answers.mkdir()
    (answers / "01-test.csv").write_text(
        "Ref,Section,Subsection,Question,MoSCoW,Status,Response,Key Technologies,POC Demo,Reference\n"
        "1.1,S,Sub,First question labelled 1.1,Must,Can do today,Answer to first,TechA,DemoA,https://example.com\n"
        "5.99,S,Sub,Second question mislabelled (should be higher ref),Must,Can do today,Answer to second,TechB,DemoB,https://example.org\n",
        encoding="utf-8",
    )

    out = tmp_path / "dup_out.xlsx"
    report = build_questionnaire_xlsx(template, answers, out)

    result = openpyxl.load_workbook(out)["Questionnaire"]
    # Template Ref values preserved (both "1.1")
    assert result.cell(row=2, column=1).value == "1.1"
    assert result.cell(row=3, column=1).value == "1.1"
    # Each row gets the correct answer by question-text match
    assert result.cell(row=2, column=7).value == "Answer to first"
    assert result.cell(row=3, column=7).value == "Answer to second"
    # No Ref mismatches reported — the answer CSV's 5.99 is consumed by the
    # template's second "1.1" row via question-text match
    assert report.extra_refs == []
