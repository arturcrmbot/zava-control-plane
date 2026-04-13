import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
wb = openpyxl.load_workbook(r'c:\dev\ghcp sdk stuff\RFP\MSFT_Response\WPP Agentic RFP - Response plan.xlsx')
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f'=== Sheet: {sheet} ===')
    for row in ws.iter_rows(values_only=True):
        vals = [str(v) if v else '' for v in row]
        if any(v for v in vals):
            print(' | '.join(vals))
    print()
