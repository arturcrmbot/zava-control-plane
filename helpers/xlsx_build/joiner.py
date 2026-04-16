"""Join the 29 answer CSVs onto WPP's original questionnaire template xlsx."""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from helpers.csv_loader import load_answer_csvs, AnswerRow


ANSWER_COLUMNS = ["Status", "Response", "Key Technologies", "POC Demo", "Reference"]

_WS = re.compile(r"\s+")


def _normalize_question(text: str) -> str:
    return _WS.sub(" ", text or "").strip()


@dataclass
class JoinReport:
    row_count: int = 0
    missing_refs: list[str] = field(default_factory=list)
    extra_refs: list[str] = field(default_factory=list)
    question_text_mismatches: list[tuple[str, str, str]] = field(default_factory=list)


def _copy_sheet(src_ws, dst_ws) -> None:
    for row in src_ws.iter_rows(values_only=False):
        for cell in row:
            dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)


def build_questionnaire_xlsx(
    template_xlsx: Path,
    answers_dir: Path,
    output_xlsx: Path,
) -> JoinReport:
    template_xlsx = Path(template_xlsx)
    answers_dir = Path(answers_dir)
    output_xlsx = Path(output_xlsx)

    answers: list[AnswerRow] = load_answer_csvs(answers_dir)
    answers_by_ref: dict[str, AnswerRow] = {a.ref: a for a in answers}

    src_wb = openpyxl.load_workbook(template_xlsx)
    try:
        if "Questionnaire" not in src_wb.sheetnames:
            raise ValueError(f"Template {template_xlsx.name} is missing the 'Questionnaire' sheet")

        out_wb = openpyxl.Workbook()
        out_wb.remove(out_wb.active)

        # Instructions sheet copied as-is
        if "Instructions" in src_wb.sheetnames:
            dst = out_wb.create_sheet("Instructions")
            _copy_sheet(src_wb["Instructions"], dst)

        q_src = src_wb["Questionnaire"]
        q_dst = out_wb.create_sheet("Questionnaire")

        src_header = [cell.value for cell in q_src[1]]
        header = src_header + ANSWER_COLUMNS
        q_dst.append(header)

        report = JoinReport()
        consumed_answer_refs: set[str] = set()

        # First pass — count occurrences of each template Ref to detect duplicates
        template_ref_counts: dict[str, int] = {}
        for row in q_src.iter_rows(min_row=2, values_only=True):
            ref = str(row[0]).strip() if row[0] else ""
            if ref:
                template_ref_counts[ref] = template_ref_counts.get(ref, 0) + 1

        # Build an index of answers by question text for duplicate-Ref matching
        answers_by_question: dict[str, AnswerRow] = {}
        for a in answers:
            answers_by_question[_normalize_question(a.question)] = a

        for row in q_src.iter_rows(min_row=2, values_only=True):
            ref = str(row[0]).strip() if row[0] else ""
            section, subsection, question, moscow = row[1], row[2], row[3], row[4]
            if not ref:
                continue

            answer = None
            if template_ref_counts[ref] > 1:
                # Duplicate Ref in template — match by question text
                answer = answers_by_question.get(_normalize_question(str(question)))
            else:
                answer = answers_by_ref.get(ref)

            if answer is None and template_ref_counts[ref] > 1:
                # We expected a question-text match; surface the miss so the operator sees it
                report.question_text_mismatches.append(
                    (ref, str(question), "NO_MATCH_FOR_DUPLICATE_REF")
                )

            if answer is None:
                report.missing_refs.append(ref)
                q_dst.append([ref, section, subsection, question, moscow, "", "", "", "", ""])
                continue

            consumed_answer_refs.add(answer.ref)

            if answer.question and question and answer.question.strip() != str(question).strip():
                report.question_text_mismatches.append((ref, str(question), answer.question))

            reference_rendered = answer.reference.replace(" | ", "\n")
            q_dst.append([
                ref, section, subsection, question, moscow,
                answer.status, answer.response, answer.key_technologies,
                answer.poc_demo, reference_rendered,
            ])
            report.row_count += 1

        # Extras — answers that weren't consumed
        for ref in answers_by_ref:
            if ref not in consumed_answer_refs:
                report.extra_refs.append(ref)

        # Formatting: wrap text, widen key columns
        for col_idx, name in enumerate(header, start=1):
            letter = get_column_letter(col_idx)
            if name == "Response":
                q_dst.column_dimensions[letter].width = 60
            elif name == "Reference":
                q_dst.column_dimensions[letter].width = 40
            elif name == "Question":
                q_dst.column_dimensions[letter].width = 50
            else:
                q_dst.column_dimensions[letter].width = 18

        wrap = Alignment(wrap_text=True, vertical="top")
        for row_cells in q_dst.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = wrap

        out_wb.save(output_xlsx)
    finally:
        src_wb.close()
    return report
